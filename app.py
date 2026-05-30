"""
谷团打表一体化工具
"""
from __future__ import annotations

import io
import math
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Tuple, Any

import json
import requests
from datetime import date
from dateutil.relativedelta import relativedelta

import pandas as pd
import streamlit as st
import xlsxwriter

try:
    from pypinyin import lazy_pinyin
    _HAS_PYPINYIN = True
except Exception:
    _HAS_PYPINYIN = False

# ─────────────────────────────────────────────
# 1. 通用工具函数与默认文件自检
# ─────────────────────────────────────────────

def get_group_key(name: str) -> str:
    name = (name or "").strip()
    if not name:
        return "#"
    if _HAS_PYPINYIN:
        try:
            p = lazy_pinyin(name)
            if p and p[0]:
                c = p[0][0].upper()
                if "A" <= c <= "Z":
                    return c
        except Exception:
            pass
    c = name[0].upper()
    return c if "A" <= c <= "Z" else "#"

def round_up(amount: float) -> float:
    return math.ceil(amount * 100) / 100

def to_num(x) -> float | None:
    try:
        return float(str(x).replace(",", "").replace("¥", "").replace("￥", "").strip())
    except Exception:
        return None

def clean_title(title: str) -> str:
    if not title:
        return ""
    return re.sub(r"\s*制表时间：.*$", "", str(title)).strip()

def get_image_bytes(img_file) -> io.BytesIO | None:
    if not img_file:
        return None
    try:
        img_file.seek(0)
        data = img_file.read()
        img_file.seek(0)
        return io.BytesIO(data)
    except Exception:
        return None

def load_default_images() -> Tuple[Any, List[Any]]:
    """自动读取仓库中预设的图片资源作为默认收款码和背景"""
    left_img = None
    qr_images = []
    
    # 自动检索默认背景
    for ext in ["png", "jpg", "jpeg"]:
        p = Path(f"left_bg.{ext}")
        if p.exists():
            try:
                with open(p, "rb") as f:
                    left_img = io.BytesIO(f.read())
                break
            except Exception:
                pass
                
    # 同时兼容 "qr_codes" 和 "收款码" 两个文件夹名字
    for folder_name in ["qr_codes", "收款码"]:
        qr_folder = Path(folder_name)
        if qr_folder.exists() and qr_folder.is_dir():
            for ext in ["png", "jpg", "jpeg", "JPG", "JPEG", "PNG"]:
                for p in sorted(qr_folder.glob(f"*.{ext}")):
                    try:
                        with open(p, "rb") as f:
                            qr_images.append(io.BytesIO(f.read()))
                    except Exception:
                        pass
            if qr_images:
                break
                
    return left_img, qr_images

# ─────────────────────────────────────────────
# 2. 源表解析与砍配过滤逻辑
# ─────────────────────────────────────────────

def parse_source_file(uploaded_file, sheet_name: str | int | None = None) -> Tuple[str, List[str], List[float | None], List[Dict]]:
    uploaded_file.seek(0)
    df = pd.read_excel(uploaded_file, sheet_name=sheet_name if sheet_name is not None else 0, header=None, engine="openpyxl", keep_default_na=True)
    header_idx = 0
    for i in range(min(8, len(df))):
        v = str(df.iat[i, 0]) if not pd.isna(df.iat[i, 0]) else ""
        if v.strip() == "种类":
            header_idx = i
            break
    
    raw_title = str(df.iat[0, 0]) if not pd.isna(df.iat[0, 0]) else ""
    title = clean_title(raw_title)
    
    # 如果表格内无标题，则默认使用工作表名称作为标题
    if not title and isinstance(sheet_name, str):
        title = sheet_name
    
    price_row_idx = header_idx + 1
    products = [str(v) for v in df.iloc[header_idx].fillna("").tolist()[1:]]
    prices = [to_num(x) for x in df.iloc[price_row_idx].fillna("").tolist()[1:]]
    
    is_single = "单领" in title
    is_box_col = ["抱盒" in p or "端盒" in p for p in products]
    details = []

    for r in range(price_row_idx + 1, len(df)):
        row = df.iloc[r].tolist()
        if pd.isna(row[0]):
            continue
        cells = row[1: 1 + len(products)]
        
        def cell_has_cn(c):
            return not (pd.isna(c) or str(c).strip() == "")
        
        if is_single:
            for j, cell in enumerate(cells):
                if cell_has_cn(cell) and j < len(products):
                    name = str(cell).strip()
                    price = prices[j]
                    if name and price and price > 0:
                        details.append({"name": name, "product": products[j], "price": price, "title": title})
        else:
            non_box_filled = [cell_has_cn(cells[j]) for j in range(len(cells)) if j < len(is_box_col) and not is_box_col[j]]
            row_ready = len(non_box_filled) > 0 and all(non_box_filled)
            for j, cell in enumerate(cells):
                if not cell_has_cn(cell) or j >= len(products):
                    continue
                if is_box_col[j] or row_ready:
                    name = str(cell).strip()
                    price = prices[j]
                    if name and price and price > 0:
                        details.append({"name": name, "product": products[j], "price": price, "title": title})
    return title, products, prices, details

def get_allocated_source_df(uploaded_file, sheet_name: str | int | None = None) -> pd.DataFrame:
    """提取经过砍配逻辑过滤之后的源表 DataFrame，未成功匹配的行与冷门单元格将被清空"""
    uploaded_file.seek(0)
    df = pd.read_excel(uploaded_file, sheet_name=sheet_name if sheet_name is not None else 0, header=None, engine="openpyxl", keep_default_na=True)
    header_idx = 0
    for i in range(min(8, len(df))):
        v = str(df.iat[i, 0]) if not pd.isna(df.iat[i, 0]) else ""
        if v.strip() == "种类":
            header_idx = i
            break
            
    raw_title = str(df.iat[0, 0]) if not pd.isna(df.iat[0, 0]) else ""
    title = clean_title(raw_title)
    
    if not title and isinstance(sheet_name, str):
        title = sheet_name
        
    price_row_idx = header_idx + 1
    products = [str(v) for v in df.iloc[header_idx].fillna("").tolist()[1:]]
    prices = [to_num(x) for x in df.iloc[price_row_idx].fillna("").tolist()[1:]]
    
    is_single = "单领" in title
    is_box_col = ["抱盒" in p or "端盒" in p for p in products]
    
    keep_rows = list(range(price_row_idx + 1))
    
    for r in range(price_row_idx + 1, len(df)):
        row = df.iloc[r].tolist()
        if pd.isna(row[0]):
            continue
        cells = row[1: 1 + len(products)]
        
        def cell_has_cn(c):
            return not (pd.isna(c) or str(c).strip() == "")
            
        if is_single:
            has_any = any(cell_has_cn(cells[j]) and j < len(products) for j in range(len(cells)))
            if has_any:
                keep_rows.append(r)
        else:
            non_box_filled = [cell_has_cn(cells[j]) for j in range(len(cells)) if j < len(is_box_col) and not is_box_col[j]]
            row_ready = len(non_box_filled) > 0 and all(non_box_filled)
            has_box_filled = any(cell_has_cn(cells[j]) for j, is_box in enumerate(is_box_col) if is_box and j < len(cells))
            
            if row_ready or has_box_filled:
                keep_rows.append(r)
                if not row_ready:
                    # 如果这盒没能拼成，只保留抱盒或端盒，其余散件清空
                    for j in range(len(products)):
                        if j < len(is_box_col) and not is_box_col[j]:
                            df.iat[r, 1 + j] = None
                            
    filtered_df = df.iloc[keep_rows].copy()
    return filtered_df

def aggregate(details: List[Dict]):
    totals = defaultdict(float)
    per_person = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for d in details:
        totals[d["name"]] += d["price"]
        per_person[d["name"]][d["title"]][d["product"]].append(d["price"])
    return per_person, totals

# ─────────────────────────────────────────────
# 3. 样式与视觉组件
# ─────────────────────────────────────────────

THEMES = {
    "红色": {"base": "#D32F2F", "light": "#FFEBEE", "group": "#FFCDD2", "font": "white"},
    "蓝色": {"base": "#1976D2", "light": "#E3F2FD", "group": "#BBDEFB", "font": "white"},
    "绿色": {"base": "#388E3C", "light": "#E8F5E9", "group": "#C8E6C9", "font": "white"},
    "浅蓝": {"base": "#00BCD4", "light": "#E0F7FA", "group": "#B2EBF2", "font": "white"},
    "黄色": {"base": "#FBC02D", "light": "#FFFDE7", "group": "#FFF9C4", "font": "#333333"},
    "橙色": {"base": "#FF8C00", "light": "#FFF5EE", "group": "#FFE6CC", "font": "white"},
    "银色": {"base": "#757575", "light": "#F5F5F5", "group": "#E0E0E0", "font": "white"},
    "粉色": {"base": "#C8336F", "light": "#FCE4EC", "group": "#F8BBD0", "font": "white"},
    "紫色": {"base": "#7B1FA2", "light": "#F3E5F5", "group": "#E1BEE7", "font": "white"},
}

def make_formats(wb: xlsxwriter.Workbook, theme_name: str) -> Dict[str, Any]:
    cfg = THEMES.get(theme_name, THEMES["橙色"])
    def f(**kw):
        base = {"font_name": "微软雅黑", "valign": "vcenter", "align": "center", "border": 1, "border_color": "#A0A0A0"}
        base.update(kw)
        return wb.add_format(base)
    return {
        "head": f(bold=True, bg_color=cfg["base"], font_color=cfg["font"]),
        "title_top": f(bold=True, bg_color=cfg["base"], font_color=cfg["font"], font_size=22),
        "ddl": f(bold=True, font_color="#D32F2F", font_size=32, border=1, align="center", text_wrap=True),
        "notice_center": f(font_size=15, bold=True, border=1, align="center", text_wrap=True),
        "left_name": f(bold=True, bg_color=cfg["light"], font_size=16, align="center", text_wrap=True),
        "odd": f(bg_color="#FFFFFF"),
        "even": f(bg_color=cfg["light"]),
        "group": f(bold=True, bg_color=cfg["group"], font_color="#000000", align="center"),
        "money_odd": f(bg_color="#FFFFFF", num_format="#,##0.00"),
        "money_even": f(bg_color=cfg["light"], num_format="#,##0.00"),
        "gu_merge": f(bold=True, bg_color=cfg["base"], font_color=cfg["font"]),
        "prod_head": f(bold=True, bg_color=cfg["light"], font_size=10),
    }

def draw_merged_group(ws, start_row, col, data_list, fmt):
    if not data_list:
        return
    last_g, start_idx = None, 0
    for i, item in enumerate(data_list):
        curr_g = item["group"] if isinstance(item, dict) else get_group_key(item)
        if curr_g != last_g:
            if i > start_idx:
                if (i - start_idx) > 1:
                    ws.merge_range(start_row + start_idx, col, start_row + i - 1, col, last_g, fmt)
                else:
                    ws.write(start_row + start_idx, col, last_g, fmt)
            start_idx, last_g = i, curr_g
    if (len(data_list) - start_idx) > 1:
        ws.merge_range(start_row + start_idx, col, start_row + len(data_list) - 1, col, last_g, fmt)
    else:
        ws.write(start_row + start_idx, col, last_g, fmt)

# ─────────────────────────────────────────────
# 4. 顶栏 1/4 四分天下自适应逻辑
# ─────────────────────────────────────────────

def draw_top_banner(ws, fmt, total_cols, title_display, notice_lines, ddl_text, left_img, qr_images):
    if total_cols < 4:
        total_cols = 4
        
    c1 = total_cols // 4
    c2 = total_cols // 2
    c3 = (total_cols * 3) // 4
    
    if c1 < 1:
        c1 = 1
    if c2 <= c1:
        c2 = c1 + 1
    if c3 <= c2:
        c3 = c2 + 1
    if total_cols <= c3:
        total_cols = c3 + 1
    
    ws.merge_range(0, 0, 0, total_cols - 1, title_display, fmt["title_top"])
    ws.set_row(0, 40)
    
    ws.merge_range(1, 0, 5, c1 - 1, title_display, fmt["left_name"])
    if left_img:
        img_data = get_image_bytes(left_img)
        if img_data:
            try:
                ws.insert_image(1, 0, "left.png", {
                    "image_data": img_data,
                    "x_scale": 0.5, "y_scale": 0.5,
                    "x_offset": 10, "y_offset": 5
                })
            except Exception:
                pass
            
    ws.merge_range(1, c1, 5, c2 - 1, "\n".join(notice_lines), fmt["notice_center"])
    
    ws.merge_range(1, c2, 5, c3 - 1, "", fmt["notice_center"])
    if qr_images:
        for idx, qr in enumerate(qr_images[:2]):
            img_data = get_image_bytes(qr)
            if img_data:
                try:
                    x_off = 20 + (idx * 120)
                    ws.insert_image(1, c2, f"qr_{idx}.png", {
                        "image_data": img_data,
                        "x_scale": 0.5, "y_scale": 0.5,
                        "x_offset": x_off, "y_offset": 15
                    })
                except Exception:
                    pass
                
    ws.merge_range(1, c3, 5, total_cols - 1, f"DDL\n{ddl_text}", fmt["ddl"])
    
    has_images = bool(left_img or qr_images)
    row_height = 55 if has_images else 22
    for r in range(1, 6):
        ws.set_row(r, row_height)

# ─────────────────────────────────────────────
# 5. 详情表写入
# ─────────────────────────────────────────────

def write_detail_sheet(ws, fmt, per_person, totals, all_titles, title_products, title_prices, rows_per_block, ddl_text, notice_lines, left_img, qr_images, custom_title):
    TOP_ROWS = 6
    col_map = [(t, p) for t in all_titles for p in title_products.get(t, [])]
    N_SIDE, N_PROD = 3, len(col_map)
    TOTAL_COLS = N_SIDE + N_PROD + N_SIDE
    sorted_names = sorted(totals.keys(), key=lambda n: (get_group_key(n), n))
    
    draw_top_banner(ws, fmt, TOTAL_COLS, custom_title or " / ".join(all_titles), notice_lines, ddl_text, left_img, qr_images)

    data_row = TOP_ROWS
    for i in range(0, len(sorted_names), rows_per_block):
        chunk = sorted_names[i : i + rows_per_block]
        for c, h in enumerate(["组", "CN", "总金额"]):
            ws.merge_range(data_row, c, data_row + 2, c, h, fmt["head"])
        curr = N_SIDE
        for t in all_titles:
            prods = title_products[t]
            n = len(prods)
            if n > 1:
                ws.merge_range(data_row, curr, data_row, curr + n - 1, t, fmt["gu_merge"])
            else:
                ws.write(data_row, curr, t, fmt["gu_merge"])
            for j, p in enumerate(prods):
                ws.write(data_row + 1, curr + j, p, fmt["prod_head"])
                price = (title_prices.get(t) or {}).get(p)
                if price is not None:
                    ws.write(data_row + 2, curr + j, price, fmt["prod_head"])
                else:
                    ws.write(data_row + 2, curr + j, "", fmt["prod_head"])
            curr += n
        for c, h in enumerate(["总金额", "CN", "组"]):
            ws.merge_range(data_row, curr + c, data_row + 2, curr + c, h, fmt["head"])
        
        data_row += 3
        for r_idx, name in enumerate(chunk):
            r = data_row + r_idx
            st_f = fmt["odd"] if r_idx % 2 == 0 else fmt["even"]
            mn_f = fmt["money_odd"] if r_idx % 2 == 0 else fmt["money_even"]
            ws.write(r, 1, name, st_f)
            ws.write(r, 2, totals[name], mn_f)
            for c_idx, (t, p) in enumerate(col_map):
                qty = len(per_person[name][t][p])
                ws.write(r, N_SIDE + c_idx, qty if qty > 0 else "", st_f)
            ws.write(r, N_SIDE + N_PROD, totals[name], mn_f)
            ws.write(r, N_SIDE + N_PROD + 1, name, st_f)
            ws.set_row(r, 18)
        
        draw_merged_group(ws, data_row, 0, chunk, fmt["group"])
        draw_merged_group(ws, data_row, TOTAL_COLS - 1, chunk, fmt["group"])
        data_row += len(chunk)

    ws.set_column(0, 0, 5)
    ws.set_column(1, 1, 22)
    ws.set_column(2, 2, 12)
    for i in range(N_PROD):
        ws.set_column(N_SIDE + i, N_SIDE + i, 5)
    ws.set_column(TOTAL_COLS - 3, TOTAL_COLS - 3, 12)
    ws.set_column(TOTAL_COLS - 2, TOTAL_COLS - 2, 22)
    ws.set_column(TOTAL_COLS - 1, TOTAL_COLS - 1, 5)

# ─────────────────────────────────────────────
# 6. 省流表写入
# ─────────────────────────────────────────────

def write_simple_sheet(ws, fmt, totals, title_text, rows_per_col, qr_images, notice_lines, ddl_text, left_img):
    TOP_ROWS = 6
    names = sorted(totals.keys(), key=lambda n: (get_group_key(n), n))
    rows_data = [{"group": get_group_key(n), "name": n, "amount": totals[n]} for n in names]
    num_blocks = math.ceil(len(rows_data) / rows_per_col)
    TOTAL_DATA_COLS = num_blocks * 4
    
    if TOTAL_DATA_COLS < 8:
        TOTAL_DATA_COLS = 8
    
    draw_top_banner(ws, fmt, TOTAL_DATA_COLS, title_text, notice_lines, ddl_text, left_img, qr_images)

    for b in range(num_blocks):
        c0 = b * 4
        chunk = rows_data[b * rows_per_col : (b + 1) * rows_per_col]
        ws.write(TOP_ROWS, c0, "组", fmt["head"])
        ws.write(TOP_ROWS, c0 + 1, "CN", fmt["head"])
        ws.write(TOP_ROWS, c0 + 2, "金额", fmt["head"])
        for i, d in enumerate(chunk):
            r = TOP_ROWS + 1 + i
            st_f = fmt["odd"] if i % 2 == 0 else fmt["even"]
            mn_f = fmt["money_odd"] if i % 2 == 0 else fmt["money_even"]
            ws.write(r, c0 + 1, d["name"], st_f)
            ws.write(r, c0 + 2, d["amount"], mn_f)
        
        draw_merged_group(ws, TOP_ROWS + 1, c0, chunk, fmt["group"])
        ws.set_column(c0, c0, 5)
        ws.set_column(c0 + 1, c0 + 1, 22) 
        ws.set_column(c0 + 2, c0 + 2, 12)
        ws.set_column(c0 + 3, c0 + 3, 1)

    for c in range(num_blocks * 4, TOTAL_DATA_COLS):
        ws.set_column(c, c, 12)

# ─────────────────────────────────────────────
# 7. 国际运费表写入
# ─────────────────────────────────────────────

def write_shipping_sheet(ws, fmt, ship_blocks, rows_per_col, title_text, ddl_text, notice_lines, left_img, qr_images):
    TOP_ROWS = 6

    all_names_set = {}
    for blk in ship_blocks:
        for e in blk["entries"]:
            n = e["name"]
            if n not in all_names_set:
                all_names_set[n] = defaultdict(lambda: defaultdict(float))
            for prod, qty in e["prod_amounts"].items():
                all_names_set[n][blk["title"]][prod] += qty

    col_map = [(blk["title"], p) for blk in ship_blocks for p in blk["products"]]
    prod_fee_map = {(blk["title"], p): blk["prod_fees"].get(p, 0.0)
                    for blk in ship_blocks for p in blk["products"]}

    sorted_names = sorted(all_names_set.keys(), key=lambda n: (get_group_key(n), n))
    rows_data = []
    for name in sorted_names:
        total_fee = sum(
            all_names_set[name][t][p] * prod_fee_map.get((t, p), 0.0)
            for t, p in col_map
        )
        rows_data.append({
            "group": get_group_key(name),
            "name": name,
            "amounts": {(t, p): all_names_set[name][t][p] for t, p in col_map},
            "total": total_fee,
        })

    if not rows_data:
        return

    N_PROD = len(col_map)
    BLOCK_W = 2 + N_PROD + 1   
    STEP = BLOCK_W + 1          
    num_blocks = math.ceil(len(rows_data) / rows_per_col)
    TOTAL_DATA_COLS = num_blocks * STEP
    if TOTAL_DATA_COLS < 8:
        TOTAL_DATA_COLS = 8

    draw_top_banner(ws, fmt, TOTAL_DATA_COLS, title_text, notice_lines, ddl_text, left_img, qr_images)

    HEADER_ROWS = 3
    data_header_row = TOP_ROWS

    for b in range(num_blocks):
        c0 = b * STEP
        chunk = rows_data[b * rows_per_col: (b + 1) * rows_per_col]

        ws.merge_range(data_header_row, c0, data_header_row + 2, c0, "组", fmt["head"])
        ws.merge_range(data_header_row, c0 + 1, data_header_row + 2, c0 + 1, "CN", fmt["head"])

        curr = c0 + 2
        for blk in ship_blocks:
            n = len(blk["products"])
            if n > 1:
                ws.merge_range(data_header_row, curr, data_header_row, curr + n - 1,
                               blk["title"], fmt["gu_merge"])
            else:
                ws.write(data_header_row, curr, blk["title"], fmt["gu_merge"])
            for j, p in enumerate(blk["products"]):
                ws.write(data_header_row + 1, curr + j, p, fmt["prod_head"])
                fee = blk["prod_fees"].get(p, 0.0)
                ws.write(data_header_row + 2, curr + j, f"¥{fee:.2f}", fmt["prod_head"])
            curr += n

        ws.merge_range(data_header_row, curr, data_header_row + 2, curr, "总运费", fmt["head"])

        data_row = data_header_row + HEADER_ROWS
        for i, d in enumerate(chunk):
            r = data_row + i
            st_f = fmt["odd"] if i % 2 == 0 else fmt["even"]
            mn_f = fmt["money_odd"] if i % 2 == 0 else fmt["money_even"]
            ws.write(r, c0 + 1, d["name"], st_f)
            for ci, (t, p) in enumerate(col_map):
                qty = d["amounts"].get((t, p), 0.0)
                ws.write(r, c0 + 2 + ci, int(qty) if qty > 0 else "", st_f)
            ws.write(r, c0 + 2 + N_PROD, d["total"], mn_f)
            ws.set_row(r, 18)

        draw_merged_group(ws, data_row, c0, chunk, fmt["group"])

        ws.set_column(c0, c0, 5)
        ws.set_column(c0 + 1, c0 + 1, 22)
        for ci in range(N_PROD):
            ws.set_column(c0 + 2 + ci, c0 + 2 + ci, 8)
        ws.set_column(c0 + 2 + N_PROD, c0 + 2 + N_PROD, 12)
        ws.set_column(c0 + BLOCK_W, c0 + BLOCK_W, 1)

# ─────────────────────────────────────────────
# 8. 其余完整函数与导入库核心提取逻辑
# ─────────────────────────────────────────────

def extract_simple_sheet(file):
    import openpyxl
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = None
    for sname in wb.sheetnames:
        if any(x in sname for x in ["省流", "汇总", "运费"]):
            ws = wb[sname]
            break
    if not ws:
        ws = wb.active
    data = {}
    for step in [4, 5]:
        for cs in range(1, ws.max_column + 1, step):
            nc, ac = cs + 1, cs + 2
            if ac > ws.max_column:
                break
            header = str(ws.cell(1, nc).value or "").lower()
            if not any(x in header for x in ["cn", "姓名", "名字"]):
                continue
            for r in range(2, ws.max_row + 1):
                n = ws.cell(r, nc).value
                a = ws.cell(r, ac).value
                if n and a:
                    try:
                        data[str(n).strip()] = data.get(str(n).strip(), 0) + float(a)
                    except Exception:
                        pass
        if data:
            break
    return data

def extract_goods_name(title: str) -> str:
    """提取货品名称，过滤嵌套括号标签与交易属性"""
    if not title:
        return ""
    
    # 1. 自动过滤尾部的制表时间
    title = clean_title(title)
    
    # 2. 自动剥离常见的排表后缀字样
    title = re.sub(r"(排表|打表|肾表|统计表|结算表|汇总表|表)$", "", title).strip()
    
    # 3. 剥离可能存在的最外层嵌套大括号或方括号
    name = title
    if name.startswith("【") and name.endswith("】"):
        name = name[1:-1].strip()
    elif name.startswith("[") and name.endswith("]"):
        name = name[1:-1].strip()
        
    # 4. 循环剥离头部的各种模式修饰前缀标签（单领/拼箱/端盒/抱盒等属性标签）
    tag_pat = r"^(【(单领|非单领|拼箱|端盒|抱盒|拼套|拼盒|定金|全款|预售|现货|代购|拼/抱|拼)】|\[(单领|非单领|拼箱|端盒|抱盒|拼套|拼盒|定金|全款|预售|现货|代购|拼/抱|拼)\])\s*"
    while True:
        prev = name
        name = re.sub(tag_pat, "", name).strip()
        if name == prev:
            break
            
    # 如果剥离到最后成空，退回到原始标题
    return name if name else title


def ai_fill_fields(goods_name: str, api_key: str) -> Tuple[str, str]:
    try:
        resp = requests.post(
            "https://api.deepseek.com/v1/chat/completions",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={
                "model": "deepseek-chat",
                "max_tokens": 100,
                "messages": [{
                    "role": "user",
                    "content": (
                        f"货品名称：{goods_name}\n"
                        "请根据货品名称推断：\n"
                        "1. 次名：货品名称的简短别称，2-6个字\n"
                        "2. 类型：从[吧唧、挂件、立牌、生写、色纸、亚克力、周边、其他]中选一个\n"
                        "只返回JSON，格式：{\"次名\": \"xxx\", \"类型\": \"xxx\"}"
                    )
                }]
            },
            timeout=15
        )
        resp.raise_for_status()
        text = resp.json()["choices"][0]["message"]["content"].strip()
        text = re.sub(r"```json|```", "", text).strip()
        obj = json.loads(text)
        return obj.get("次名", ""), obj.get("类型", "")
    except Exception:
        return "", ""


def build_import_records(selected_inputs: list[Tuple[Any, str]], api_key: str) -> Tuple[pd.DataFrame, List[str]]:
    """从用户自主选择的排表 Sheet 记录中提取并分析导入库数据"""
    today = date.today()
    stock_deadline = today + relativedelta(months=4)
    drop_deadline = stock_deadline + relativedelta(months=1)
    stock_str = stock_deadline.strftime("%Y/%m/%d")
    drop_str = drop_deadline.strftime("%Y/%m/%d")

    logs: List[str] = []
    rows: List[Dict] = []

    for f, sheet_name in selected_inputs:
        f.seek(0)
        try:
            title, products, prices, details = parse_source_file(f, sheet_name=sheet_name)
        except Exception as e:
            logs.append(f"{f.name} - {sheet_name} 解析失败: {e}")
            continue

        # 默认提取货品名称（使用优化后的解包算法）
        goods_name = extract_goods_name(title) if title else sheet_name

        # AI 辅助填写次名和类型
        ci_name, ci_type = "", ""
        if api_key:
            ci_name, ci_type = ai_fill_fields(goods_name, api_key)
            if not ci_name or not ci_type:
                logs.append(f"「{goods_name}」AI 返回为空，次名/类型留空")
        else:
            logs.append("未填写 API Key，次名/类型留空")

        # 统计每个 CN 购买的每个产品数量
        cn_prod: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
        for d in details:
            cn_prod[d["name"]][d["product"]] += 1

        # 每个 CN × 每个产品 生成一行导入记录
        for cn, prod_qtys in cn_prod.items():
            for prod, qty in prod_qtys.items():
                rows.append({
                    "货品名称": goods_name,
                    "次名": ci_name,
                    "人物": prod,
                    "类型": ci_type,
                    "数量": qty,
                    "CN": cn,
                    "物品状态": "未到货",
                    "囤货期限": stock_str,
                    "掉落期限": drop_str,
                })

    df = pd.DataFrame(rows, columns=["货品名称","次名","人物","类型","数量","CN","物品状态","囤货期限","掉落期限"])
    return df, logs


def write_import_sheet(wb, df, fmt):
    ws = wb.add_worksheet("导入库")
    fds = ["货品名称", "次名", "人物", "类型", "数量", "CN", "物品状态", "囤货期限", "掉落期限"]
    col_widths = [18, 12, 10, 10, 6, 16, 10, 14, 14]
    for i, h in enumerate(fds):
        ws.write(0, i, h, fmt["head"])
        ws.set_column(i, i, col_widths[i])
    ws.set_row(0, 18)
    for r_idx, row in df.iterrows():
        r = r_idx + 1
        st_f = fmt["odd"] if r_idx % 2 == 0 else fmt["even"]
        for c_idx, fd in enumerate(fds):
            ws.write(r, c_idx, str(row[fd]), st_f)
        ws.set_row(r, 16)

def write_source_sheet(wb, title, uploaded_file, used_names, sheet_name=None):
    """写入源表：自动应用砍配逻辑，只导出保留配上的有效数据"""
    try:
        df = get_allocated_source_df(uploaded_file, sheet_name=sheet_name)
        
        nm = re.sub(r"[\\/*\[\]:?'\"<>|]", "-", title)[:28] or "源表"
        bs, sf = nm, 1
        while nm in used_names:
            nm = f"{bs}({sf})"
            sf += 1
        used_names.add(nm)
        ws = wb.add_worksheet(nm)
        
        for r_idx, row in enumerate(df.values):
            for c_idx, val in enumerate(row):
                if pd.notna(val):
                    ws.write(r_idx, c_idx, val)
    except Exception:
        pass

# ═══════════════════════════════════════════════
# 9. Streamlit UI
# ═══════════════════════════════════════════════

st.set_page_config(page_title="谷团工具箱", layout="wide")

default_left, default_qrs = load_default_images()

with st.sidebar:
    st.header("全局顶栏设置")
    theme_choice = st.selectbox("选择主题色", list(THEMES.keys()), index=5)
    st.divider()
    global_title = st.text_input("表格标题", value="")
    global_ddl = st.text_input("付款 DDL ", value="无特殊情况填写打表当日向后第七个自然日的晚上22:00")
    global_notice = st.text_area("中心备注大字 (中左 25%)", value="1.转账备注cn+本期谷子\n2.蓝绿双通无手续费\n3.拖肾一周内，每天请补交1r手续费，一周后未交掉落。两次及以上掉落记录飞机票\n4....", height=180)
    
    global_left_img = st.file_uploader("左侧背景图（不传则读取预设）", type=["png","jpg","jpeg"])
    if not global_left_img and default_left:
        st.caption("💡 已自动检索到默认左侧背景图 (left_bg)")
        chosen_left_img = default_left
    else:
        chosen_left_img = global_left_img

    global_qr_imgs = st.file_uploader("收款码（可多选，不传则读取预设）", type=["png","jpg","jpeg"], accept_multiple_files=True)
    if not global_qr_imgs and default_qrs:
        st.caption(f"💡 已自动检索到默认收款码文件夹中的 {len(default_qrs)} 张收款码 (qr_codes/)")
        chosen_qr_imgs = default_qrs
    else:
        chosen_qr_imgs = global_qr_imgs

    st.divider()
    rows_per_col_global = st.number_input("省流表每栏行数", 10, 100, 25)
    rows_per_block_global = st.number_input("详情表重复表头行数", 10, 100, 30)

st.title("谷团工具箱")
tab1, tab2, tab3, tab4 = st.tabs(["国际运费", "谷子肾表", "退补款提取", "导入库"])

with tab1:
    st.subheader("国际运费计算")
    ship_files = st.file_uploader("上传排表 (运费)", type=["xlsx"], accept_multiple_files=True)
    sc1, sc2 = st.columns(2)
    with sc1:
        scheme = st.radio("计费方案", ("方案一：总运费均摊", "方案二：克单价直算"), horizontal=True)
    with sc2:
        val_ship = st.number_input("金额（方案一填总运费，方案二填克单价）", 0.0, format="%.4f")

    if ship_files:
        all_vcols = {}
        for f in ship_files:
            f.seek(0)
            xl = pd.ExcelFile(f)
            sheet_names = xl.sheet_names
            
            default_exclude = ["详情表", "省流表", "汇总表", "运费表", "导入库", "Sheet1"]
            candidate_sheets = [s for s in sheet_names if s not in default_exclude]
            if not candidate_sheets:
                candidate_sheets = sheet_names
                
            selected_sheets = st.multiselect(
                f"选择文件【{f.name}】中参与计算运费的 Sheet (可自由增删)",
                options=sheet_names,
                default=candidate_sheets,
                key=f"sheets_{f.name}"
            )
            
            if not selected_sheets:
                continue
                
            for sheet in selected_sheets:
                df_s_all = pd.read_excel(f, sheet_name=sheet, header=None, engine="openpyxl")
                header_idx = 0
                for i in range(min(8, len(df_s_all))):
                    v = str(df_s_all.iat[i, 0]) if not pd.isna(df_s_all.iat[i, 0]) else ""
                    if v.strip() == "种类":
                        header_idx = i
                        break
                
                raw_title = str(df_s_all.iat[0, 0]) if not pd.isna(df_s_all.iat[0, 0]) else sheet
                clean_t = clean_title(raw_title) if raw_title else sheet
                
                f.seek(0)
                df_s = pd.read_excel(f, sheet_name=sheet, header=header_idx, engine="openpyxl")
                bad = ["种类","序号","单极","总计","合计","Unnamed","备注","补款","余量"]
                cols = [c for c in df_s.columns if str(c).strip() and not any(b in str(c) for b in bad)]
                
                valid = []
                for c in cols:
                    vals = df_s[c].iloc[1:].dropna().astype(str).str.strip()
                    has_cn = vals[~vals.str.match(r"^\d+\.?\d*$")].any()
                    if has_cn:
                        valid.append(c)
                
                if valid:
                    all_vcols[f"{f.name} - {sheet}"] = {
                        "cols": valid, 
                        "df": df_s, 
                        "title": clean_t,
                        "file_name": f.name,
                        "sheet_name": sheet
                    }

        for fn, info in all_vcols.items():
            ky = f"ship_{fn}"
            
            with st.expander(f"配置重量: 【{info['title']}】 排表 (工作表: {info['sheet_name']})"):
                
                # ── 新的模式选择：完全贴合你的逻辑 ──
                mode_key = f"mode_{ky}"
                if mode_key not in st.session_state:
                    st.session_state[mode_key] = "A. 整盒配比"
                    
                mode = st.radio(
                    "模式",
                    ["A. 整盒配比", "B. 混合/关键词", "C. 统一重量"],
                    key=mode_key,
                    horizontal=True
                )
                
                if mode == "A. 整盒配比":
                    col1, col2, col3 = st.columns([2, 2, 2])
                    with col1:
                        bt = st.number_input("整盒重", key=f"bt_{ky}", min_value=0.0, step=10.0, format="%.2f")
                    with col2:
                        bc = st.number_input("数量", key=f"bc_{ky}", value=1, min_value=1)
                    with col3:
                        box_kw = st.text_input("端盒关键词", key=f"box_kw_{ky}", value="端盒,抱盒")
                    
                    if st.button("计算填充", key=f"btn_calc_{ky}"):
                        kws = [k.strip() for k in box_kw.replace("，", ",").split(",") if k.strip()]
                        avg_val = bt / max(bc, 1)
                        for c in info["cols"]:
                            w_key = f"inp_{ky}_{c}"
                            is_box = any(kw in str(c) for kw in kws)
                            st.session_state[w_key] = bt if is_box else avg_val
                        st.rerun()
                        
                elif mode == "B. 混合/关键词":
                    col1, col2, col3 = st.columns([4, 2, 2])
                    with col1:
                        kw_str = st.text_input("关键词", key=f"kw_{ky}")
                    with col2:
                        kw_val = st.number_input("重量", key=f"kw_val_{ky}", min_value=0.0, step=1.0, format="%.2f")
                    with col3:
                        st.write("") # 占位
                        st.write("") # 占位
                        if st.button("应用", key=f"btn_kw_{ky}", use_container_width=True):
                            if kw_str.strip():
                                for c in info["cols"]:
                                    if kw_str.strip() in str(c):
                                        w_key = f"inp_{ky}_{c}"
                                        st.session_state[w_key] = kw_val
                                st.rerun()
                            else:
                                st.warning("请输入关键词后再点击应用")
                                
                elif mode == "C. 统一重量":
                    col1, col2 = st.columns([4, 2])
                    with col1:
                        uni_val = st.number_input("统一重量", key=f"uni_val_{ky}", min_value=0.0, step=1.0, format="%.2f")
                    with col2:
                        st.write("") # 占位
                        st.write("") # 占位
                        if st.button("应用", key=f"btn_uni_{ky}", use_container_width=True):
                            for c in info["cols"]:
                                w_key = f"inp_{ky}_{c}"
                                st.session_state[w_key] = uni_val
                            st.rerun()
                            
                st.divider()
                
                # ── 模块 4：手工微调网格 ──
                st.subheader("手工调整各单品重量 (g)")
                ci = st.columns(4)
                for i, c in enumerate(info["cols"]):
                    w_key = f"inp_{ky}_{c}"
                    if w_key not in st.session_state:
                        st.session_state[w_key] = 0.0
                        
                    # ⚠️ 修复关键：不使用赋值形式。
                    # 在 Streamlit 中，将输入组件直接绑定到 key 时，无需用“等号”重新把返回值赋回 session_state。
                    ci[i % 4].number_input(
                        str(c),
                        key=w_key,
                        min_value=0.0,
                        step=1.0,
                        format="%.2f"
                    )

        if st.button("生成运费表", type="primary"):
            ud: Dict[str, Dict] = {}
            tw = 0.0
            
            # 第一步：汇总所有人购买的件数和对应的总重量
            for fn, info in all_vcols.items():
                ky = f"ship_{fn}"
                for col in info["cols"]:
                    w_key = f"inp_{ky}_{col}"
                    w = st.session_state.get(w_key, 0.0)
                    for vn in info["df"][col].iloc[1:]:
                        if pd.notna(vn):
                            n = str(vn).strip()
                            if not n or n.lower() in ["nan", "none"]:
                                continue
                            if n not in ud:
                                ud[n] = {"w": 0.0, "prod_qtys": defaultdict(lambda: defaultdict(float))}
                            ud[n]["w"] += w
                            ud[n]["prod_qtys"][info["title"]][str(col)] += 1
                            tw += w

            if tw == 0:
                st.warning("当前各单品配置的总重量为0，请先配置重量。")
            else:
                # 第二步：生成各个产品列的运费单价
                ship_blocks = []
                for fn, info in all_vcols.items():
                    ky = f"ship_{fn}"
                    prods = info["cols"]
                    pf = {}
                    for col in prods:
                        w_key = f"inp_{ky}_{col}"
                        w = st.session_state.get(w_key, 0.0)
                        if "一" in scheme:
                            pf[col] = round_up(w * val_ship / tw) if tw > 0 else 0.0
                        else:
                            pf[col] = round_up(w * val_ship)

                    entries = []
                    for name, data in ud.items():
                        prod_amounts = {col: data["prod_qtys"].get(info["title"], {}).get(col, 0)
                                        for col in prods}
                        if any(v > 0 for v in prod_amounts.values()):
                            entries.append({
                                "group": get_group_key(name),
                                "name": name,
                                "prod_amounts": prod_amounts,
                            })
                    ship_blocks.append({
                        "title": info["title"],
                        "products": prods,
                        "prod_fees": pf,
                        "entries": entries,
                    })

                out = io.BytesIO()
                wb_s = xlsxwriter.Workbook(out)
                fmt_s = make_formats(wb_s, theme_choice)
                write_shipping_sheet(
                    wb_s.add_worksheet("运费表"), fmt_s, ship_blocks,
                    rows_per_col_global,
                    global_title or "国际运费",
                    global_ddl, global_notice.splitlines(),
                    chosen_left_img, chosen_qr_imgs
                )
                wb_s.close()
                st.success(f"运费表生成完成！共 {len(ud)} 人，总重 {tw:.1f}g")
                
                download_name = f"{global_title or '运费表'}_总表.xlsx"
                st.download_button("下载运费表", out.getvalue(), download_name)

with tab2:
    gu_files = st.file_uploader("上传排表 (肾表)", type=["xlsx"], accept_multiple_files=True)
    if gu_files and st.button("生成对称合并肾表", type="primary"):
        all_d, all_t, t_prods, t_prices, src_d = [], [], {}, {}, []
        for f in gu_files:
            f.seek(0)
            t, p, pr, d = parse_source_file(f)
            if t not in all_t:
                all_t.append(t)
            t_prods[t] = p
            t_prices[t] = {p[i]: pr[i] for i in range(len(p)) if i < len(pr) and pr[i] is not None}
            all_d.extend(d)
            f.seek(0)
            src_d.append((t, f))
        pp, tt = aggregate(all_d)

        out = io.BytesIO()
        wb_out = xlsxwriter.Workbook(out)
        fmt = make_formats(wb_out, theme_choice)
        write_detail_sheet(wb_out.add_worksheet("详情表"), fmt, pp, tt, all_t, t_prods, t_prices,
                           rows_per_block_global, global_ddl, global_notice.splitlines(),
                           chosen_left_img, chosen_qr_imgs, global_title)
        write_simple_sheet(wb_out.add_worksheet("省流表"), fmt, tt,
                           global_title or " / ".join(all_t),
                           rows_per_col_global, chosen_qr_imgs,
                           global_notice.splitlines(), global_ddl, chosen_left_img)
        usd = {"详情表", "省流表"}
        
        for t, f in src_d:
            f.seek(0)
            write_source_sheet(wb_out, t, f, usd)
            
        wb_out.close()
        st.success("肾表生成完成！源工作表数据已过滤至砍配之后的分配结果。")
        
        download_name = f"{global_title or '肾表'}_总表.xlsx"
        st.download_button("下载肾表", out.getvalue(), download_name)

with tab3:
    st.subheader("提取差额")
    f_or = st.file_uploader("上传原表 (xlsx)", type=["xlsx"])
    f_ad = st.file_uploader("上传新表 (xlsx)", type=["xlsx"])
    if f_or and f_ad and st.button("计算退补差额"):
        d1 = extract_simple_sheet(f_or)
        d2 = extract_simple_sheet(f_ad)
        res = [{"name": n, "diff": d2.get(n,0)-d1.get(n,0)} for n in set(d1)|set(d2) if d2.get(n,0)-d1.get(n,0) != 0]
        st.dataframe(pd.DataFrame(res))

with tab4:
    st.subheader("生成导入库表")
    st.caption("从原始排表提取数据，AI自动填写次名和类型")
    im_files = st.file_uploader("上传排表 xlsx（可多选）", type=["xlsx"], accept_multiple_files=True)
    deepseek_key = st.text_input("DeepSeek API Key（不填则次名/类型留空）", type="password")

    # 汇总自主选择的 Sheet
    selected_im_inputs = []
    if im_files:
        for f in im_files:
            f.seek(0)
            xl = pd.ExcelFile(f)
            sheet_names = xl.sheet_names
            
            # 自动过滤汇总表和省流表
            default_exclude = ["详情表", "省流表", "汇总表", "运费表", "导入库", "Sheet1"]
            candidate_sheets = [s for s in sheet_names if s not in default_exclude]
            if not candidate_sheets:
                candidate_sheets = sheet_names
                
            selected_sheets = st.multiselect(
                f"选择文件【{f.name}】中导入库存的 Sheet (可自由删减)",
                options=sheet_names,
                default=candidate_sheets,
                key=f"im_sheets_{f.name}"
            )
            
            for s in selected_sheets:
                selected_im_inputs.append((f, s))

    if im_files and st.button("🚀 生成导入库表", type="primary"):
        if not selected_im_inputs:
            st.warning("请至少选择一个工作表 (Sheet) 进行导入库计算！")
        else:
            with st.spinner("处理中，AI 推断字段可能需要几秒..."):
                df_im, logs = build_import_records(selected_im_inputs, deepseek_key)
            if logs:
                with st.expander("⚠️ 处理日志"):
                    for l in logs:
                        st.caption(l)
            if df_im.empty:
                st.warning("未提取到任何记录，请检查文件格式")
            else:
                st.success(f"提取到 {len(df_im)} 条记录")
                st.dataframe(df_im, use_container_width=True)
                out_im = io.BytesIO()
                wb_im = xlsxwriter.Workbook(out_im)
                fmt_im = make_formats(wb_im, theme_choice)
                write_import_sheet(wb_im, df_im, fmt_im)
                wb_im.close()
                st.download_button("下载导入库表", out_im.getvalue(), "导入库.xlsx")
